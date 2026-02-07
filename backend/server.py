from fastapi import FastAPI, APIRouter, HTTPException, Depends, Response, UploadFile, File, BackgroundTasks
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import asyncio
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook, load_workbook
import resend

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

resend.api_key = os.environ.get('RESEND_API_KEY', '')
ALERT_EMAIL = os.environ.get('ALERT_EMAIL', '')
ALERT_DAYS_BEFORE = int(os.environ.get('ALERT_DAYS_BEFORE', 7))
SENDER_EMAIL = "onboarding@resend.dev"

app = FastAPI()
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

JWT_SECRET = os.environ.get('JWT_SECRET', 'warehouse-construction-secret-key-2024')
JWT_ALGORITHM = 'HS256'

UPLOAD_DIR = ROOT_DIR / 'uploads'
UPLOAD_DIR.mkdir(exist_ok=True)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ==================== AUTH MODELS ====================
class UserCreate(BaseModel):
    name: str
    email: EmailStr
    password: str

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    email: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserResponse

# ==================== EQUIPAMENTO MODEL ====================
class EquipamentoCreate(BaseModel):
    codigo: str
    descricao: str
    marca: str = ""
    modelo: str = ""
    data_aquisicao: Optional[str] = None
    ativo: bool = True
    categoria: str = ""
    numero_serie: str = ""
    estado_conservacao: str = "Bom"
    foto: str = ""
    obra_id: Optional[str] = None
    # Novos campos
    manual_url: str = ""
    certificado_url: str = ""
    ficha_manutencao_url: str = ""
    em_manutencao: bool = False
    descricao_avaria: str = ""

class Equipamento(EquipamentoCreate):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tipo: str = "Equipamento"
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# ==================== VIATURA MODEL ====================
class ViaturaCreate(BaseModel):
    matricula: str
    marca: str = ""
    modelo: str = ""
    combustivel: str = "Gasoleo"
    ativa: bool = True
    foto: str = ""
    data_vistoria: Optional[str] = None
    data_seguro: Optional[str] = None
    documento_unico: str = ""
    apolice_seguro: str = ""
    observacoes: str = ""
    obra_id: Optional[str] = None
    # Novos campos - Documentação
    dua_url: str = ""
    seguro_url: str = ""
    ipo_url: str = ""
    carta_verde_url: str = ""
    manual_url: str = ""
    # Novos campos - Manutenção
    em_manutencao: bool = False
    descricao_avaria: str = ""
    # Novos campos - Datas para alertas
    data_ipo: Optional[str] = None
    data_proxima_revisao: Optional[str] = None
    kms_atual: int = 0
    kms_proxima_revisao: int = 0

class Viatura(ViaturaCreate):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# ==================== MATERIAL MODEL ====================
class MaterialCreate(BaseModel):
    codigo: str
    descricao: str
    unidade: str = "unidade"
    stock_atual: float = 0
    stock_minimo: float = 0
    ativo: bool = True

class Material(MaterialCreate):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# ==================== OBRA MODEL ====================
class ObraCreate(BaseModel):
    codigo: str
    nome: str
    endereco: str = ""
    cliente: str = ""
    estado: str = "Ativa"

class Obra(ObraCreate):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# ==================== MOVIMENTO MODEL ====================
class MovimentoCreate(BaseModel):
    recurso_id: str
    tipo_recurso: str  # equipamento, viatura
    tipo_movimento: str  # Saida, Devolucao
    obra_id: Optional[str] = None
    responsavel_levantou: str = ""
    responsavel_devolveu: str = ""
    data_levantamento: Optional[str] = None
    data_devolucao: Optional[str] = None
    observacoes: str = ""

class AtribuirRecursoRequest(BaseModel):
    recurso_id: str
    tipo_recurso: str  # equipamento, viatura
    obra_id: str
    responsavel_levantou: str = ""
    data_levantamento: Optional[str] = None
    observacoes: str = ""

class DevolverRecursoRequest(BaseModel):
    recurso_id: str
    tipo_recurso: str  # equipamento, viatura
    responsavel_devolveu: str = ""
    data_devolucao: Optional[str] = None
    observacoes: str = ""

class Movimento(MovimentoCreate):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# ==================== MOVIMENTO STOCK MODEL ====================
class MovimentoStockCreate(BaseModel):
    material_id: str
    tipo_movimento: str
    quantidade: float
    obra_id: Optional[str] = None
    fornecedor: str = ""
    documento: str = ""
    responsavel: str = ""
    observacoes: str = ""

class MovimentoStock(MovimentoStockCreate):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    data_hora: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# ==================== MOVIMENTO VIATURA MODEL ====================
class MovimentoViaturaCreate(BaseModel):
    viatura_id: str
    obra_id: Optional[str] = None
    condutor: str = ""
    km_inicial: float = 0
    km_final: float = 0
    data: str = ""
    observacoes: str = ""

class MovimentoViatura(MovimentoViaturaCreate):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# ==================== AUTH FUNCTIONS ====================
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode(), hashed.encode())

def create_token(user_id: str) -> str:
    payload = {
        "sub": user_id,
        "exp": datetime.now(timezone.utc) + timedelta(days=7)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# ==================== EMAIL FUNCTIONS ====================
async def send_alert_email(alerts):
    if not ALERT_EMAIL or not resend.api_key or not alerts:
        return
    
    alerts_html = ""
    for alert in alerts:
        tipo = "Vistoria" if alert["tipo_alerta"] == "vistoria" else "Seguro"
        urgency_class = "color: #dc2626;" if alert["dias_restantes"] <= 0 else "color: #f97316;"
        status = "EXPIRADO" if alert["dias_restantes"] <= 0 else f"Expira em {alert['dias_restantes']} dias"
        
        alerts_html += f"""
        <tr>
            <td style="padding: 12px; border-bottom: 1px solid #333;">{alert['matricula']}</td>
            <td style="padding: 12px; border-bottom: 1px solid #333;">{alert['marca']} {alert['modelo']}</td>
            <td style="padding: 12px; border-bottom: 1px solid #333;">{tipo}</td>
            <td style="padding: 12px; border-bottom: 1px solid #333;">{alert['data_expiracao']}</td>
            <td style="padding: 12px; border-bottom: 1px solid #333; {urgency_class} font-weight: bold;">{status}</td>
        </tr>
        """
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <body style="font-family: Arial, sans-serif; margin: 0; padding: 20px; background-color: #1a1a1a; color: #fff;">
        <div style="max-width: 600px; margin: 0 auto; background-color: #2a2a2a; border-radius: 4px; overflow: hidden;">
            <div style="background-color: #f97316; color: #000; padding: 20px; text-align: center;">
                <h1 style="margin: 0; font-size: 24px;">⚠️ Alertas de Viaturas</h1>
                <p style="margin: 10px 0 0 0;">José Firmino - Construção Civil</p>
            </div>
            <div style="padding: 20px;">
                <p style="color: #ccc; margin-bottom: 20px;">
                    As seguintes viaturas têm vistorias ou seguros a expirar nos próximos {ALERT_DAYS_BEFORE} dias:
                </p>
                <table style="width: 100%; border-collapse: collapse; font-size: 14px; color: #fff;">
                    <thead>
                        <tr style="background-color: #333;">
                            <th style="padding: 12px; text-align: left;">Matrícula</th>
                            <th style="padding: 12px; text-align: left;">Viatura</th>
                            <th style="padding: 12px; text-align: left;">Tipo</th>
                            <th style="padding: 12px; text-align: left;">Data</th>
                            <th style="padding: 12px; text-align: left;">Estado</th>
                        </tr>
                    </thead>
                    <tbody>{alerts_html}</tbody>
                </table>
            </div>
        </div>
    </body>
    </html>
    """
    
    try:
        email = await asyncio.to_thread(resend.Emails.send, {
            "from": SENDER_EMAIL,
            "to": [ALERT_EMAIL],
            "subject": f"⚠️ Alertas de Viaturas - {len(alerts)} alerta(s)",
            "html": html_content
        })
        return email
    except Exception as e:
        logger.error(f"Failed to send alert email: {str(e)}")
        raise

# ==================== AUTH ROUTES ====================
@api_router.post("/auth/register", response_model=TokenResponse)
async def register(data: UserCreate):
    existing = await db.users.find_one({"email": data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_id = str(uuid.uuid4())
    user_doc = {
        "id": user_id,
        "name": data.name,
        "email": data.email,
        "password": hash_password(data.password),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    
    token = create_token(user_id)
    return TokenResponse(access_token=token, user=UserResponse(id=user_id, name=data.name, email=data.email))

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(data: UserLogin):
    user = await db.users.find_one({"email": data.email}, {"_id": 0})
    if not user or not verify_password(data.password, user["password"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    token = create_token(user["id"])
    return TokenResponse(access_token=token, user=UserResponse(id=user["id"], name=user["name"], email=user["email"]))

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(user=Depends(get_current_user)):
    return UserResponse(id=user["id"], name=user["name"], email=user["email"])

# ==================== UPLOAD ROUTES ====================
@api_router.post("/upload")
async def upload_file(file: UploadFile = File(...), user=Depends(get_current_user)):
    if not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Only image files are allowed")
    
    ext = file.filename.split(".")[-1] if "." in file.filename else "jpg"
    filename = f"{uuid.uuid4()}.{ext}"
    filepath = UPLOAD_DIR / filename
    
    content = await file.read()
    with open(filepath, "wb") as f:
        f.write(content)
    
    return {"url": f"/api/uploads/{filename}", "filename": filename}

@api_router.post("/upload/pdf")
async def upload_pdf(file: UploadFile = File(...), user=Depends(get_current_user)):
    """Upload PDF documents (manuals, certificates, etc.)"""
    if file.content_type != "application/pdf":
        raise HTTPException(status_code=400, detail="Apenas ficheiros PDF são permitidos")
    
    # Max 10MB
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Ficheiro demasiado grande (máx. 10MB)")
    
    filename = f"{uuid.uuid4()}.pdf"
    filepath = UPLOAD_DIR / filename
    
    with open(filepath, "wb") as f:
        f.write(content)
    
    return {"url": f"/api/uploads/{filename}", "filename": filename, "original_name": file.filename}

@api_router.get("/uploads/{filename}")
async def get_upload(filename: str):
    filepath = UPLOAD_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    ext = filename.split(".")[-1].lower()
    content_types = {
        "jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png", 
        "gif": "image/gif", "webp": "image/webp", "pdf": "application/pdf"
    }
    
    with open(filepath, "rb") as f:
        content = f.read()
    
    return Response(content=content, media_type=content_types.get(ext, "application/octet-stream"))

# ==================== EQUIPAMENTO ROUTES ====================
@api_router.get("/equipamentos")
async def get_equipamentos(user=Depends(get_current_user)):
    items = await db.equipamentos.find({}, {"_id": 0}).to_list(1000)
    # Garantir que os campos novos têm valores por defeito
    for item in items:
        item.setdefault("em_manutencao", False)
        item.setdefault("descricao_avaria", "")
        item.setdefault("manual_url", "")
        item.setdefault("certificado_url", "")
        item.setdefault("ficha_manutencao_url", "")
    return items

@api_router.get("/equipamentos/{equipamento_id}")
async def get_equipamento(equipamento_id: str, user=Depends(get_current_user)):
    item = await db.equipamentos.find_one({"id": equipamento_id}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="Equipamento não encontrado")
    
    # Garantir que os campos novos têm valores por defeito
    item.setdefault("em_manutencao", False)
    item.setdefault("descricao_avaria", "")
    item.setdefault("manual_url", "")
    item.setdefault("certificado_url", "")
    item.setdefault("ficha_manutencao_url", "")
    
    # Get obra info if assigned
    obra = None
    if item.get("obra_id"):
        obra = await db.obras.find_one({"id": item["obra_id"]}, {"_id": 0})
    
    # Get movement history
    movimentos = await db.movimentos.find(
        {"recurso_id": equipamento_id, "tipo_recurso": "equipamento"}, 
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    # Enrich movements with obra name
    for mov in movimentos:
        if mov.get("obra_id"):
            obra_mov = await db.obras.find_one({"id": mov["obra_id"]}, {"_id": 0, "nome": 1, "codigo": 1})
            if obra_mov:
                mov["obra_nome"] = obra_mov.get("nome", "")
                mov["obra_codigo"] = obra_mov.get("codigo", "")
    
    return {"equipamento": item, "obra_atual": obra, "historico": movimentos}

class ManutencaoUpdate(BaseModel):
    em_manutencao: bool
    descricao_avaria: str = ""

@api_router.patch("/equipamentos/{equipamento_id}/manutencao")
async def update_equipamento_manutencao(equipamento_id: str, data: ManutencaoUpdate, user=Depends(get_current_user)):
    """Atualizar estado de manutenção de um equipamento (sem editar outros campos)"""
    existing = await db.equipamentos.find_one({"id": equipamento_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Equipamento não encontrado")
    
    update_data = {"em_manutencao": data.em_manutencao, "descricao_avaria": data.descricao_avaria}
    await db.equipamentos.update_one({"id": equipamento_id}, {"$set": update_data})
    
    updated = await db.equipamentos.find_one({"id": equipamento_id}, {"_id": 0})
    return updated

@api_router.post("/equipamentos")
async def create_equipamento(data: EquipamentoCreate, user=Depends(get_current_user)):
    existing = await db.equipamentos.find_one({"codigo": data.codigo}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Código já existe")
    
    equipamento = Equipamento(**data.model_dump())
    await db.equipamentos.insert_one(equipamento.model_dump())
    return equipamento

@api_router.put("/equipamentos/{equipamento_id}")
async def update_equipamento(equipamento_id: str, data: EquipamentoCreate, user=Depends(get_current_user)):
    existing = await db.equipamentos.find_one({"id": equipamento_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Equipamento não encontrado")
    
    await db.equipamentos.update_one({"id": equipamento_id}, {"$set": data.model_dump()})
    return await db.equipamentos.find_one({"id": equipamento_id}, {"_id": 0})

@api_router.delete("/equipamentos/{equipamento_id}")
async def delete_equipamento(equipamento_id: str, user=Depends(get_current_user)):
    result = await db.equipamentos.delete_one({"id": equipamento_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Equipamento não encontrado")
    return {"message": "Equipamento eliminado"}

# ==================== VIATURA ROUTES ====================
def set_viatura_defaults(item):
    """Garantir valores por defeito nos campos novos"""
    item.setdefault("em_manutencao", False)
    item.setdefault("descricao_avaria", "")
    item.setdefault("dua_url", "")
    item.setdefault("seguro_url", "")
    item.setdefault("ipo_url", "")
    item.setdefault("carta_verde_url", "")
    item.setdefault("manual_url", "")
    item.setdefault("data_ipo", None)
    item.setdefault("data_proxima_revisao", None)
    item.setdefault("kms_atual", 0)
    item.setdefault("kms_proxima_revisao", 0)
    return item

@api_router.get("/viaturas")
async def get_viaturas(user=Depends(get_current_user)):
    items = await db.viaturas.find({}, {"_id": 0}).to_list(1000)
    for item in items:
        set_viatura_defaults(item)
    return items

@api_router.get("/viaturas/{viatura_id}")
async def get_viatura(viatura_id: str, user=Depends(get_current_user)):
    item = await db.viaturas.find_one({"id": viatura_id}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="Viatura não encontrada")
    
    set_viatura_defaults(item)
    
    obra = None
    if item.get("obra_id"):
        obra = await db.obras.find_one({"id": item["obra_id"]}, {"_id": 0})
    
    movimentos = await db.movimentos.find(
        {"recurso_id": viatura_id, "tipo_recurso": "viatura"}, 
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    # Enrich movements with obra name
    for mov in movimentos:
        if mov.get("obra_id"):
            obra_mov = await db.obras.find_one({"id": mov["obra_id"]}, {"_id": 0, "nome": 1, "codigo": 1})
            if obra_mov:
                mov["obra_nome"] = obra_mov.get("nome", "")
                mov["obra_codigo"] = obra_mov.get("codigo", "")
    
    km_movimentos = await db.movimentos_viaturas.find(
        {"viatura_id": viatura_id}, {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    # Calcular alertas
    alertas = []
    from datetime import datetime, timedelta
    hoje = datetime.now()
    
    # Alerta seguro
    if item.get("data_seguro"):
        try:
            data_seg = datetime.fromisoformat(item["data_seguro"].replace("Z", "+00:00"))
            dias_seguro = (data_seg - hoje).days
            if dias_seguro <= 30:
                alertas.append({"tipo": "seguro", "mensagem": f"Seguro expira em {dias_seguro} dias", "urgente": dias_seguro <= 7})
        except: pass
    
    # Alerta IPO
    if item.get("data_ipo"):
        try:
            data_ipo = datetime.fromisoformat(item["data_ipo"].replace("Z", "+00:00"))
            dias_ipo = (data_ipo - hoje).days
            if dias_ipo <= 30:
                alertas.append({"tipo": "ipo", "mensagem": f"IPO expira em {dias_ipo} dias", "urgente": dias_ipo <= 7})
        except: pass
    
    # Alerta revisão por data
    if item.get("data_proxima_revisao"):
        try:
            data_rev = datetime.fromisoformat(item["data_proxima_revisao"].replace("Z", "+00:00"))
            dias_rev = (data_rev - hoje).days
            if dias_rev <= 30:
                alertas.append({"tipo": "revisao", "mensagem": f"Revisão em {dias_rev} dias", "urgente": dias_rev <= 7})
        except: pass
    
    # Alerta revisão por KMs
    if item.get("kms_proxima_revisao") and item.get("kms_atual"):
        kms_faltam = item["kms_proxima_revisao"] - item["kms_atual"]
        if kms_faltam <= 1000 and kms_faltam > 0:
            alertas.append({"tipo": "kms", "mensagem": f"Revisão em {kms_faltam} km", "urgente": kms_faltam <= 500})
    
    return {"viatura": item, "obra_atual": obra, "historico": movimentos, "km_historico": km_movimentos, "alertas": alertas}

@api_router.patch("/viaturas/{viatura_id}/manutencao")
async def update_viatura_manutencao(viatura_id: str, data: ManutencaoUpdate, user=Depends(get_current_user)):
    """Atualizar estado de manutenção de uma viatura (sem editar outros campos)"""
    existing = await db.viaturas.find_one({"id": viatura_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Viatura não encontrada")
    
    update_data = {"em_manutencao": data.em_manutencao, "descricao_avaria": data.descricao_avaria}
    await db.viaturas.update_one({"id": viatura_id}, {"$set": update_data})
    
    updated = await db.viaturas.find_one({"id": viatura_id}, {"_id": 0})
    return updated

@api_router.post("/viaturas")
async def create_viatura(data: ViaturaCreate, user=Depends(get_current_user)):
    existing = await db.viaturas.find_one({"matricula": data.matricula}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Matrícula já existe")
    
    viatura = Viatura(**data.model_dump())
    await db.viaturas.insert_one(viatura.model_dump())
    return viatura

@api_router.put("/viaturas/{viatura_id}")
async def update_viatura(viatura_id: str, data: ViaturaCreate, user=Depends(get_current_user)):
    existing = await db.viaturas.find_one({"id": viatura_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Viatura não encontrada")
    
    await db.viaturas.update_one({"id": viatura_id}, {"$set": data.model_dump()})
    return await db.viaturas.find_one({"id": viatura_id}, {"_id": 0})

@api_router.delete("/viaturas/{viatura_id}")
async def delete_viatura(viatura_id: str, user=Depends(get_current_user)):
    result = await db.viaturas.delete_one({"id": viatura_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Viatura não encontrada")
    return {"message": "Viatura eliminada"}

# ==================== MATERIAL ROUTES ====================
@api_router.get("/materiais")
async def get_materiais(user=Depends(get_current_user)):
    return await db.materiais.find({}, {"_id": 0}).to_list(1000)

@api_router.post("/materiais")
async def create_material(data: MaterialCreate, user=Depends(get_current_user)):
    existing = await db.materiais.find_one({"codigo": data.codigo}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Código já existe")
    
    material = Material(**data.model_dump())
    await db.materiais.insert_one(material.model_dump())
    return material

@api_router.put("/materiais/{material_id}")
async def update_material(material_id: str, data: MaterialCreate, user=Depends(get_current_user)):
    existing = await db.materiais.find_one({"id": material_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Material não encontrado")
    
    await db.materiais.update_one({"id": material_id}, {"$set": data.model_dump()})
    return await db.materiais.find_one({"id": material_id}, {"_id": 0})

@api_router.get("/materiais/{material_id}")
async def get_material_detail(material_id: str, user=Depends(get_current_user)):
    """Get material with movement history"""
    material = await db.materiais.find_one({"id": material_id}, {"_id": 0})
    if not material:
        raise HTTPException(status_code=404, detail="Material não encontrado")
    
    # Get movement history
    historico = await db.movimentos_stock.find(
        {"material_id": material_id}, 
        {"_id": 0}
    ).sort("data_hora", -1).to_list(100)
    
    return {"material": material, "historico": historico}

@api_router.delete("/materiais/{material_id}")
async def delete_material(material_id: str, user=Depends(get_current_user)):
    result = await db.materiais.delete_one({"id": material_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Material não encontrado")
    return {"message": "Material eliminado"}

# ==================== OBRA ROUTES ====================
@api_router.get("/obras")
async def get_obras(user=Depends(get_current_user)):
    return await db.obras.find({}, {"_id": 0}).to_list(1000)

@api_router.get("/obras/{obra_id}")
async def get_obra(obra_id: str, user=Depends(get_current_user)):
    obra = await db.obras.find_one({"id": obra_id}, {"_id": 0})
    if not obra:
        raise HTTPException(status_code=404, detail="Obra não encontrada")
    
    equipamentos = await db.equipamentos.find({"obra_id": obra_id}, {"_id": 0}).to_list(1000)
    viaturas = await db.viaturas.find({"obra_id": obra_id}, {"_id": 0}).to_list(1000)
    
    return {"obra": obra, "equipamentos": equipamentos, "viaturas": viaturas}

@api_router.post("/obras")
async def create_obra(data: ObraCreate, user=Depends(get_current_user)):
    existing = await db.obras.find_one({"codigo": data.codigo}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Código já existe")
    
    obra = Obra(**data.model_dump())
    await db.obras.insert_one(obra.model_dump())
    return obra

@api_router.put("/obras/{obra_id}")
async def update_obra(obra_id: str, data: ObraCreate, user=Depends(get_current_user)):
    existing = await db.obras.find_one({"id": obra_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Obra não encontrada")
    
    await db.obras.update_one({"id": obra_id}, {"$set": data.model_dump()})
    return await db.obras.find_one({"id": obra_id}, {"_id": 0})

@api_router.delete("/obras/{obra_id}")
async def delete_obra(obra_id: str, user=Depends(get_current_user)):
    result = await db.obras.delete_one({"id": obra_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Obra não encontrada")
    
    # Remove obra association from resources
    await db.equipamentos.update_many({"obra_id": obra_id}, {"$set": {"obra_id": None}})
    await db.viaturas.update_many({"obra_id": obra_id}, {"$set": {"obra_id": None}})
    return {"message": "Obra eliminada"}

# ==================== MOVIMENTO (Atribuição) ROUTES ====================
@api_router.post("/movimentos/atribuir")
async def atribuir_recurso(data: AtribuirRecursoRequest, user=Depends(get_current_user)):
    """Atribuir equipamento ou viatura a uma obra"""
    collection = db.equipamentos if data.tipo_recurso == "equipamento" else db.viaturas
    
    recurso = await collection.find_one({"id": data.recurso_id}, {"_id": 0})
    if not recurso:
        raise HTTPException(status_code=404, detail="Recurso não encontrado")
    
    # Check if already assigned to another obra
    if recurso.get("obra_id") and recurso["obra_id"] != data.obra_id:
        obra_atual = await db.obras.find_one({"id": recurso["obra_id"]}, {"_id": 0})
        raise HTTPException(
            status_code=400, 
            detail=f"Este recurso já está atribuído à obra: {obra_atual['nome'] if obra_atual else 'Desconhecida'}"
        )
    
    # Create movement record
    movimento = Movimento(
        recurso_id=data.recurso_id,
        tipo_recurso=data.tipo_recurso,
        tipo_movimento="Saida",
        obra_id=data.obra_id,
        responsavel_levantou=data.responsavel_levantou,
        data_levantamento=data.data_levantamento or datetime.now(timezone.utc).isoformat(),
        observacoes=data.observacoes
    )
    await db.movimentos.insert_one(movimento.model_dump())
    
    # Update resource
    await collection.update_one({"id": data.recurso_id}, {"$set": {"obra_id": data.obra_id}})
    
    return {"message": "Recurso atribuído com sucesso", "movimento_id": movimento.id}

@api_router.post("/movimentos/devolver")
async def devolver_recurso(data: DevolverRecursoRequest, user=Depends(get_current_user)):
    """Devolver equipamento ou viatura de uma obra"""
    collection = db.equipamentos if data.tipo_recurso == "equipamento" else db.viaturas
    
    recurso = await collection.find_one({"id": data.recurso_id}, {"_id": 0})
    if not recurso:
        raise HTTPException(status_code=404, detail="Recurso não encontrado")
    
    # Create movement record
    movimento = Movimento(
        recurso_id=data.recurso_id,
        tipo_recurso=data.tipo_recurso,
        tipo_movimento="Devolucao",
        obra_id=recurso.get("obra_id"),
        responsavel_devolveu=data.responsavel_devolveu,
        data_devolucao=data.data_devolucao or datetime.now(timezone.utc).isoformat(),
        observacoes=data.observacoes
    )
    await db.movimentos.insert_one(movimento.model_dump())
    
    # Remove obra association
    await collection.update_one({"id": data.recurso_id}, {"$set": {"obra_id": None}})
    
    return {"message": "Recurso devolvido com sucesso", "movimento_id": movimento.id}

@api_router.get("/movimentos")
async def get_movimentos(user=Depends(get_current_user)):
    return await db.movimentos.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)

# ==================== MOVIMENTO STOCK ROUTES ====================
@api_router.get("/movimentos/stock")
async def get_movimentos_stock(user=Depends(get_current_user)):
    return await db.movimentos_stock.find({}, {"_id": 0}).to_list(1000)

@api_router.post("/movimentos/stock")
async def create_movimento_stock(data: MovimentoStockCreate, user=Depends(get_current_user)):
    movimento = MovimentoStock(**data.model_dump())
    await db.movimentos_stock.insert_one(movimento.model_dump())
    
    material = await db.materiais.find_one({"id": data.material_id}, {"_id": 0})
    if material:
        new_stock = material.get("stock_atual", 0)
        if data.tipo_movimento == "Entrada":
            new_stock += data.quantidade
        else:
            new_stock -= data.quantidade
        await db.materiais.update_one({"id": data.material_id}, {"$set": {"stock_atual": new_stock}})
    
    return movimento

# ==================== MOVIMENTO VIATURA ROUTES ====================
@api_router.get("/movimentos/viaturas")
async def get_movimentos_viaturas(user=Depends(get_current_user)):
    return await db.movimentos_viaturas.find({}, {"_id": 0}).to_list(1000)

@api_router.post("/movimentos/viaturas")
async def create_movimento_viatura(data: MovimentoViaturaCreate, user=Depends(get_current_user)):
    movimento = MovimentoViatura(**data.model_dump())
    await db.movimentos_viaturas.insert_one(movimento.model_dump())
    return movimento

# ==================== ALERTS ROUTES ====================
@api_router.get("/alerts/check")
async def check_alerts(user=Depends(get_current_user)):
    viaturas = await db.viaturas.find({"ativa": True}, {"_id": 0}).to_list(1000)
    today = datetime.now(timezone.utc).date()
    alerts = []
    
    for v in viaturas:
        for field, tipo in [("data_vistoria", "vistoria"), ("data_seguro", "seguro")]:
            if v.get(field):
                try:
                    date = datetime.fromisoformat(v[field].replace("Z", "+00:00")).date()
                    days_until = (date - today).days
                    if days_until <= ALERT_DAYS_BEFORE:
                        alerts.append({
                            "viatura_id": v["id"],
                            "matricula": v["matricula"],
                            "marca": v.get("marca", ""),
                            "modelo": v.get("modelo", ""),
                            "tipo_alerta": tipo,
                            "data_expiracao": date.strftime("%d/%m/%Y"),
                            "dias_restantes": days_until
                        })
                except:
                    pass
    
    return {"alerts": alerts, "total": len(alerts)}

@api_router.post("/alerts/send")
async def send_alerts(user=Depends(get_current_user)):
    if not ALERT_EMAIL or not resend.api_key:
        raise HTTPException(status_code=400, detail="Configuração de email incompleta")
    
    check_result = await check_alerts(user)
    alerts = check_result["alerts"]
    
    if not alerts:
        return {"status": "success", "message": "Não há alertas para enviar", "alerts_count": 0}
    
    try:
        await send_alert_email(alerts)
        return {"status": "success", "message": f"Email enviado com {len(alerts)} alerta(s)", "alerts_count": len(alerts)}
    except Exception as e:
        error_msg = str(e)
        if "verify a domain" in error_msg.lower():
            raise HTTPException(status_code=400, detail=f"Para enviar emails para {ALERT_EMAIL}, precisa de verificar o domínio em resend.com/domains")
        raise HTTPException(status_code=500, detail=f"Erro ao enviar email: {error_msg}")

# ==================== IMPORT/EXPORT ROUTES ====================
@api_router.post("/import/excel")
async def import_excel(file: UploadFile = File(...), user=Depends(get_current_user)):
    """Import data from Excel file"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Apenas ficheiros Excel são permitidos")
    
    content = await file.read()
    wb = load_workbook(BytesIO(content))
    imported = {"equipamentos": 0, "viaturas": 0, "materiais": 0, "obras": 0}
    
    # Import Equipamentos
    if "Equipamentos" in wb.sheetnames or "Equipamento" in wb.sheetnames:
        ws = wb["Equipamentos"] if "Equipamentos" in wb.sheetnames else wb["Equipamento"]
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            data = dict(zip(headers, row))
            codigo = str(data.get("Codigo", data.get("codigo", data.get("Código", ""))))
            if not codigo:
                continue
            
            existing = await db.equipamentos.find_one({"codigo": codigo})
            if existing:
                continue
            
            equipamento = Equipamento(
                codigo=codigo,
                descricao=str(data.get("Descricao", data.get("descricao", data.get("Descrição", "")))),
                marca=str(data.get("Marca", data.get("marca", "")) or ""),
                modelo=str(data.get("Modelo", data.get("modelo", "")) or ""),
                categoria=str(data.get("Categoria", data.get("categoria", "")) or ""),
                numero_serie=str(data.get("Numero_Serie", data.get("numero_serie", data.get("Nº Série", ""))) or ""),
                estado_conservacao=str(data.get("Estado_Conservacao", data.get("estado_conservacao", data.get("Estado", "Bom"))) or "Bom"),
                ativo=str(data.get("Ativo", data.get("ativo", "Sim"))).lower() in ["sim", "true", "1", "yes"]
            )
            await db.equipamentos.insert_one(equipamento.model_dump())
            imported["equipamentos"] += 1
    
    # Import Viaturas
    if "Viaturas" in wb.sheetnames or "Viatura" in wb.sheetnames:
        ws = wb["Viaturas"] if "Viaturas" in wb.sheetnames else wb["Viatura"]
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            data = dict(zip(headers, row))
            matricula = str(data.get("Matricula", data.get("matricula", data.get("Matrícula", ""))))
            if not matricula:
                continue
            
            existing = await db.viaturas.find_one({"matricula": matricula})
            if existing:
                continue
            
            viatura = Viatura(
                matricula=matricula,
                marca=str(data.get("Marca", data.get("marca", "")) or ""),
                modelo=str(data.get("Modelo", data.get("modelo", "")) or ""),
                combustivel=str(data.get("Combustivel", data.get("combustivel", data.get("Combustível", "Gasoleo"))) or "Gasoleo"),
                ativa=str(data.get("Ativa", data.get("ativa", "Sim"))).lower() in ["sim", "true", "1", "yes"]
            )
            await db.viaturas.insert_one(viatura.model_dump())
            imported["viaturas"] += 1
    
    # Import Materiais
    if "Materiais" in wb.sheetnames or "Material" in wb.sheetnames:
        ws = wb["Materiais"] if "Materiais" in wb.sheetnames else wb["Material"]
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            data = dict(zip(headers, row))
            codigo = str(data.get("Codigo", data.get("codigo", data.get("Código", "")))  or data.get("ID_Material", ""))
            if not codigo:
                continue
            
            existing = await db.materiais.find_one({"codigo": codigo})
            if existing:
                continue
            
            material = Material(
                codigo=codigo,
                descricao=str(data.get("Descricao", data.get("descricao", data.get("Descrição", ""))) or ""),
                unidade=str(data.get("Unidade", data.get("unidade", "unidade")) or "unidade"),
                stock_minimo=float(data.get("Stock_Minimo", data.get("stock_minimo", 0)) or 0)
            )
            await db.materiais.insert_one(material.model_dump())
            imported["materiais"] += 1
    
    # Import Obras
    if "Obras" in wb.sheetnames or "Obra" in wb.sheetnames:
        ws = wb["Obras"] if "Obras" in wb.sheetnames else wb["Obra"]
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            data = dict(zip(headers, row))
            codigo = str(data.get("Codigo", data.get("codigo", data.get("ID_Obra", ""))))
            if not codigo:
                continue
            
            existing = await db.obras.find_one({"codigo": codigo})
            if existing:
                continue
            
            obra = Obra(
                codigo=codigo,
                nome=str(data.get("Nome", data.get("nome", "")) or ""),
                estado=str(data.get("Estado", data.get("estado", "Ativa")) or "Ativa")
            )
            await db.obras.insert_one(obra.model_dump())
            imported["obras"] += 1
    
    return {"message": "Importação concluída", "imported": imported}

@api_router.get("/export/excel")
async def export_excel(user=Depends(get_current_user)):
    equipamentos = await db.equipamentos.find({}, {"_id": 0}).to_list(1000)
    viaturas = await db.viaturas.find({}, {"_id": 0}).to_list(1000)
    materiais = await db.materiais.find({}, {"_id": 0}).to_list(1000)
    obras = await db.obras.find({}, {"_id": 0}).to_list(1000)
    
    wb = Workbook()
    
    ws = wb.active
    ws.title = "Equipamentos"
    ws.append(["Código", "Descrição", "Marca", "Modelo", "Categoria", "Nº Série", "Estado", "Ativo"])
    for e in equipamentos:
        ws.append([e.get("codigo"), e.get("descricao"), e.get("marca"), e.get("modelo"), 
                   e.get("categoria"), e.get("numero_serie"), e.get("estado_conservacao"),
                   "Sim" if e.get("ativo") else "Não"])
    
    ws = wb.create_sheet("Viaturas")
    ws.append(["Matrícula", "Marca", "Modelo", "Combustível", "Data Vistoria", "Data Seguro", "Ativa"])
    for v in viaturas:
        ws.append([v.get("matricula"), v.get("marca"), v.get("modelo"), v.get("combustivel"),
                   v.get("data_vistoria"), v.get("data_seguro"), "Sim" if v.get("ativa") else "Não"])
    
    ws = wb.create_sheet("Materiais")
    ws.append(["Código", "Descrição", "Unidade", "Stock Atual", "Stock Mínimo", "Ativo"])
    for m in materiais:
        ws.append([m.get("codigo"), m.get("descricao"), m.get("unidade"),
                   m.get("stock_atual"), m.get("stock_minimo"), "Sim" if m.get("ativo") else "Não"])
    
    ws = wb.create_sheet("Obras")
    ws.append(["Código", "Nome", "Endereço", "Cliente", "Estado"])
    for o in obras:
        ws.append([o.get("codigo"), o.get("nome"), o.get("endereco"), o.get("cliente"), o.get("estado")])
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=dados_armazem.xlsx"}
    )

@api_router.get("/export/pdf")
async def export_pdf(user=Depends(get_current_user)):
    equipamentos = await db.equipamentos.find({}, {"_id": 0}).to_list(1000)
    viaturas = await db.viaturas.find({}, {"_id": 0}).to_list(1000)
    materiais = await db.materiais.find({}, {"_id": 0}).to_list(1000)
    obras = await db.obras.find({}, {"_id": 0}).to_list(1000)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []
    
    elements.append(Paragraph("José Firmino - Gestão de Armazém", styles['Title']))
    elements.append(Paragraph(f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    summary_data = [
        ["Categoria", "Total", "Ativos/Ativas"],
        ["Equipamentos", len(equipamentos), len([e for e in equipamentos if e.get("ativo")])],
        ["Viaturas", len(viaturas), len([v for v in viaturas if v.get("ativa")])],
        ["Materiais", len(materiais), "-"],
        ["Obras", len(obras), len([o for o in obras if o.get("estado") == "Ativa"])]
    ]
    
    table = Table(summary_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.976, 0.451, 0.086)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    return Response(content=buffer.getvalue(), media_type="application/pdf",
                    headers={"Content-Disposition": "attachment; filename=relatorio_armazem.pdf"})

# ==================== SUMMARY ROUTE ====================
@api_router.get("/summary")
async def get_summary(user=Depends(get_current_user)):
    equipamentos = await db.equipamentos.find({}, {"_id": 0}).to_list(1000)
    viaturas = await db.viaturas.find({}, {"_id": 0}).to_list(1000)
    materiais = await db.materiais.find({}, {"_id": 0}).to_list(1000)
    obras = await db.obras.find({}, {"_id": 0}).to_list(1000)
    
    alerts = []
    today = datetime.now(timezone.utc).date()
    
    for v in viaturas:
        for field, tipo, msg in [("data_vistoria", "vistoria", "Vistoria"), ("data_seguro", "seguro", "Seguro")]:
            if v.get(field):
                try:
                    date = datetime.fromisoformat(v[field].replace("Z", "+00:00")).date()
                    days_until = (date - today).days
                    if days_until <= ALERT_DAYS_BEFORE:
                        alerts.append({
                            "type": tipo,
                            "item": f"{v['marca']} {v['modelo']} ({v['matricula']})",
                            "message": f"{msg} em {days_until} dias" if days_until >= 0 else f"{msg} expirado",
                            "urgent": days_until < 0
                        })
                except:
                    pass
    
    for m in materiais:
        if m.get("stock_atual", 0) <= m.get("stock_minimo", 0) and m.get("stock_minimo", 0) > 0:
            alerts.append({
                "type": "stock",
                "item": f"{m['codigo']} - {m['descricao']}",
                "message": f"Stock baixo: {m.get('stock_atual', 0)} {m.get('unidade', 'un')}",
                "urgent": m.get("stock_atual", 0) == 0
            })
    
    return {
        "equipamentos": {
            "total": len(equipamentos),
            "ativos": len([e for e in equipamentos if e.get("ativo", True)]),
            "em_obra": len([e for e in equipamentos if e.get("obra_id")])
        },
        "viaturas": {
            "total": len(viaturas),
            "ativas": len([v for v in viaturas if v.get("ativa", True)]),
            "em_obra": len([v for v in viaturas if v.get("obra_id")])
        },
        "materiais": {
            "total": len(materiais),
            "stock_total": sum(m.get("stock_atual", 0) for m in materiais)
        },
        "obras": {
            "total": len(obras),
            "ativas": len([o for o in obras if o.get("estado") == "Ativa"])
        },
        "alerts": alerts
    }

# ==================== RELATÓRIOS AVANÇADOS ====================
@api_router.get("/relatorios/movimentos")
async def get_relatorio_movimentos(
    obra_id: Optional[str] = None,
    mes: Optional[int] = None,
    ano: Optional[int] = None,
    tipo_recurso: Optional[str] = None,
    user=Depends(get_current_user)
):
    """Relatório de movimentos de equipamentos e viaturas filtrado por obra e período"""
    # Build query filter
    query = {}
    
    if obra_id:
        query["obra_id"] = obra_id
    
    if tipo_recurso:
        query["tipo_recurso"] = tipo_recurso
    
    # Date filtering
    if mes and ano:
        start_date = datetime(ano, mes, 1, tzinfo=timezone.utc)
        if mes == 12:
            end_date = datetime(ano + 1, 1, 1, tzinfo=timezone.utc)
        else:
            end_date = datetime(ano, mes + 1, 1, tzinfo=timezone.utc)
        query["created_at"] = {"$gte": start_date.isoformat(), "$lt": end_date.isoformat()}
    elif ano:
        start_date = datetime(ano, 1, 1, tzinfo=timezone.utc)
        end_date = datetime(ano + 1, 1, 1, tzinfo=timezone.utc)
        query["created_at"] = {"$gte": start_date.isoformat(), "$lt": end_date.isoformat()}
    
    movimentos = await db.movimentos.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Enrich with resource and obra details
    enriched = []
    for mov in movimentos:
        item = {**mov}
        
        # Get resource details
        if mov.get("tipo_recurso") == "equipamento":
            recurso = await db.equipamentos.find_one({"id": mov["recurso_id"]}, {"_id": 0, "codigo": 1, "descricao": 1})
            if recurso:
                item["recurso_codigo"] = recurso.get("codigo", "")
                item["recurso_descricao"] = recurso.get("descricao", "")
        elif mov.get("tipo_recurso") == "viatura":
            recurso = await db.viaturas.find_one({"id": mov["recurso_id"]}, {"_id": 0, "matricula": 1, "marca": 1, "modelo": 1})
            if recurso:
                item["recurso_codigo"] = recurso.get("matricula", "")
                item["recurso_descricao"] = f"{recurso.get('marca', '')} {recurso.get('modelo', '')}"
        
        # Get obra details
        if mov.get("obra_id"):
            obra = await db.obras.find_one({"id": mov["obra_id"]}, {"_id": 0, "codigo": 1, "nome": 1})
            if obra:
                item["obra_codigo"] = obra.get("codigo", "")
                item["obra_nome"] = obra.get("nome", "")
        
        enriched.append(item)
    
    # Statistics
    total_saidas = len([m for m in movimentos if m.get("tipo_movimento") == "Saida"])
    total_devolucoes = len([m for m in movimentos if m.get("tipo_movimento") == "Devolucao"])
    equipamentos_movidos = len(set([m["recurso_id"] for m in movimentos if m.get("tipo_recurso") == "equipamento"]))
    viaturas_movidas = len(set([m["recurso_id"] for m in movimentos if m.get("tipo_recurso") == "viatura"]))
    
    return {
        "movimentos": enriched,
        "estatisticas": {
            "total_movimentos": len(movimentos),
            "total_saidas": total_saidas,
            "total_devolucoes": total_devolucoes,
            "equipamentos_movidos": equipamentos_movidos,
            "viaturas_movidas": viaturas_movidas
        }
    }

@api_router.get("/relatorios/stock")
async def get_relatorio_stock(
    obra_id: Optional[str] = None,
    mes: Optional[int] = None,
    ano: Optional[int] = None,
    user=Depends(get_current_user)
):
    """Relatório de movimentos de stock (materiais) filtrado por obra e período"""
    query = {}
    
    if obra_id:
        query["obra_id"] = obra_id
    
    if mes and ano:
        start_date = datetime(ano, mes, 1, tzinfo=timezone.utc)
        if mes == 12:
            end_date = datetime(ano + 1, 1, 1, tzinfo=timezone.utc)
        else:
            end_date = datetime(ano, mes + 1, 1, tzinfo=timezone.utc)
        query["data_hora"] = {"$gte": start_date.isoformat(), "$lt": end_date.isoformat()}
    elif ano:
        start_date = datetime(ano, 1, 1, tzinfo=timezone.utc)
        end_date = datetime(ano + 1, 1, 1, tzinfo=timezone.utc)
        query["data_hora"] = {"$gte": start_date.isoformat(), "$lt": end_date.isoformat()}
    
    movimentos = await db.movimentos_stock.find(query, {"_id": 0}).sort("data_hora", -1).to_list(1000)
    
    # Enrich with material and obra details
    enriched = []
    materiais_gastos = {}
    
    for mov in movimentos:
        item = {**mov}
        
        # Get material details
        material = await db.materiais.find_one({"id": mov["material_id"]}, {"_id": 0, "codigo": 1, "descricao": 1, "unidade": 1})
        if material:
            item["material_codigo"] = material.get("codigo", "")
            item["material_descricao"] = material.get("descricao", "")
            item["material_unidade"] = material.get("unidade", "un")
            
            # Track consumption by material
            mat_id = mov["material_id"]
            if mat_id not in materiais_gastos:
                materiais_gastos[mat_id] = {
                    "codigo": material.get("codigo", ""),
                    "descricao": material.get("descricao", ""),
                    "unidade": material.get("unidade", "un"),
                    "entradas": 0,
                    "saidas": 0
                }
            if mov.get("tipo_movimento") == "Entrada":
                materiais_gastos[mat_id]["entradas"] += mov.get("quantidade", 0)
            else:
                materiais_gastos[mat_id]["saidas"] += mov.get("quantidade", 0)
        
        # Get obra details
        if mov.get("obra_id"):
            obra = await db.obras.find_one({"id": mov["obra_id"]}, {"_id": 0, "codigo": 1, "nome": 1})
            if obra:
                item["obra_codigo"] = obra.get("codigo", "")
                item["obra_nome"] = obra.get("nome", "")
        
        enriched.append(item)
    
    # Statistics
    total_entradas = sum(m.get("quantidade", 0) for m in movimentos if m.get("tipo_movimento") == "Entrada")
    total_saidas = sum(m.get("quantidade", 0) for m in movimentos if m.get("tipo_movimento") == "Saida")
    
    return {
        "movimentos": enriched,
        "materiais_resumo": list(materiais_gastos.values()),
        "estatisticas": {
            "total_movimentos": len(movimentos),
            "total_entradas": total_entradas,
            "total_saidas": total_saidas,
            "consumo_liquido": total_saidas - total_entradas,
            "materiais_diferentes": len(materiais_gastos)
        }
    }

@api_router.get("/relatorios/obra/{obra_id}")
async def get_relatorio_obra(
    obra_id: str,
    mes: Optional[int] = None,
    ano: Optional[int] = None,
    user=Depends(get_current_user)
):
    """Relatório completo de uma obra específica"""
    obra = await db.obras.find_one({"id": obra_id}, {"_id": 0})
    if not obra:
        raise HTTPException(status_code=404, detail="Obra não encontrada")
    
    # Get resources currently assigned
    equipamentos_atuais = await db.equipamentos.find({"obra_id": obra_id}, {"_id": 0}).to_list(100)
    viaturas_atuais = await db.viaturas.find({"obra_id": obra_id}, {"_id": 0}).to_list(100)
    
    # Get movement history for this obra
    mov_query = {"obra_id": obra_id}
    stock_query = {"obra_id": obra_id}
    
    if mes and ano:
        start_date = datetime(ano, mes, 1, tzinfo=timezone.utc)
        if mes == 12:
            end_date = datetime(ano + 1, 1, 1, tzinfo=timezone.utc)
        else:
            end_date = datetime(ano, mes + 1, 1, tzinfo=timezone.utc)
        mov_query["created_at"] = {"$gte": start_date.isoformat(), "$lt": end_date.isoformat()}
        stock_query["data_hora"] = {"$gte": start_date.isoformat(), "$lt": end_date.isoformat()}
    elif ano:
        start_date = datetime(ano, 1, 1, tzinfo=timezone.utc)
        end_date = datetime(ano + 1, 1, 1, tzinfo=timezone.utc)
        mov_query["created_at"] = {"$gte": start_date.isoformat(), "$lt": end_date.isoformat()}
        stock_query["data_hora"] = {"$gte": start_date.isoformat(), "$lt": end_date.isoformat()}
    
    movimentos_ativos = await db.movimentos.find(mov_query, {"_id": 0}).sort("created_at", -1).to_list(500)
    movimentos_stock = await db.movimentos_stock.find(stock_query, {"_id": 0}).sort("data_hora", -1).to_list(500)
    
    # Calculate stock consumption by material
    consumo_materiais = {}
    for mov in movimentos_stock:
        mat_id = mov["material_id"]
        material = await db.materiais.find_one({"id": mat_id}, {"_id": 0, "codigo": 1, "descricao": 1, "unidade": 1})
        if material:
            if mat_id not in consumo_materiais:
                consumo_materiais[mat_id] = {
                    "codigo": material.get("codigo", ""),
                    "descricao": material.get("descricao", ""),
                    "unidade": material.get("unidade", "un"),
                    "quantidade_gasta": 0
                }
            if mov.get("tipo_movimento") == "Saida":
                consumo_materiais[mat_id]["quantidade_gasta"] += mov.get("quantidade", 0)
    
    return {
        "obra": obra,
        "recursos_atuais": {
            "equipamentos": equipamentos_atuais,
            "viaturas": viaturas_atuais
        },
        "periodo_filtrado": {
            "mes": mes,
            "ano": ano
        },
        "estatisticas": {
            "equipamentos_atuais": len(equipamentos_atuais),
            "viaturas_atuais": len(viaturas_atuais),
            "movimentos_ativos": len(movimentos_ativos),
            "movimentos_stock": len(movimentos_stock),
            "total_saidas_ativos": len([m for m in movimentos_ativos if m.get("tipo_movimento") == "Saida"]),
            "total_devolucoes": len([m for m in movimentos_ativos if m.get("tipo_movimento") == "Devolucao"])
        },
        "consumo_materiais": list(consumo_materiais.values())
    }

@api_router.get("/")
async def root():
    return {"message": "José Firmino - API de Gestão de Armazém"}

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
